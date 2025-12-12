from odoo import models, fields
from datetime import date, timedelta
import io
import base64
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class StockManagement(models.TransientModel):
    _name = "stock.management.report.wizard"
    _description = "Stock Management Report Wizard"

    start_date = fields.Date("Start Date",default=date.today() - timedelta(days=30))
    end_date = fields.Date("End Date", default=date.today())
    file_data = fields.Binary("File")
    file_name = fields.Char("File Name")
    report_type = fields.Selection(selection=[
        ('client_stock_report','Client Stock Report'),
        ('inventory_held_report','Inventory Held Report'),
        ('pfa_pension_report','PFA Pension Report'),
        ('employee_paye_report','Employee Paye Report'),
        ('payroll_schedule_report','Payroll Schedule Report')
    ])

    def download(self):
        if self.report_type == "client_stock_report":
           return self.client_stock_data()
        elif self.report_type == "inventory_held_report":
            return self.inventory_held_data()
        elif self.report_type == "pfa_pension_report":
            return self.pfa_pension_report()
        elif self.report_type == "employee_paye_report":
            return self.employee_paye_report()
        elif self.report_type == "payroll_schedule_report":
            return self._generate_payroll_schedule_report()

    # GENERATE EMPLOYEE PAYE REPORT
    def employee_paye_report(self):
        if self.start_date and self.end_date:
            query = """
                select hr_pay.id as id, hr_emp.id as emp_id, hr_emp.staff_id as staff_id, hr_emp.name as emp_name, 
                hr_emp.tax_payer_number as tax_payer_number, hr_sal_rul.code as code, hr_pay_line.amount as amount, hr_con.wage as wage
                from hr_payslip as hr_pay 
                left join hr_employee as hr_emp on hr_pay.employee_id = hr_emp.id
                left join hr_payslip_line as hr_pay_line on hr_pay.id = hr_pay_line.slip_id
                left join hr_salary_rule as hr_sal_rul on hr_pay_line.salary_rule_id = hr_sal_rul.id
                left join hr_contract as hr_con on hr_emp.id = hr_con.employee_id and hr_con.state = 'open'
                where hr_pay.state = 'done' and hr_pay.date_from >= %s and hr_pay.date_to <= %s
            """
            params = (self.start_date, self.end_date)
        else:
            query = """
                select hr_pay.id as id, hr_emp.id as emp_id, hr_emp.staff_id as staff_id, hr_emp.name as emp_name, 
                hr_emp.tax_payer_number as tax_payer_number, hr_sal_rul.code as code, hr_pay_line.amount as amount, hr_con.wage as wage
                from hr_payslip as hr_pay 
                left join hr_employee as hr_emp on hr_pay.employee_id = hr_emp.id
                left join hr_payslip_line as hr_pay_line on hr_pay.id = hr_pay_line.slip_id
                left join hr_salary_rule as hr_sal_rul on hr_pay_line.salary_rule_id = hr_sal_rul.id
                left join hr_contract as hr_con on hr_emp.id = hr_con.employee_id and hr_con.state = 'open'
                where hr_pay.state = 'done'
            """
            params = ()

        self.env.cr.execute(query, params)
        hr_payslips = self.env.cr.dictfetchall()

        employee_dict = {}

        for payslip in hr_payslips: 
            emp_id = payslip['emp_id']
            tax_payable = taxable_amount = 0
            if payslip['code'] == "GROSS":
                taxable_amount = payslip['amount']
            if payslip['code'] == "PAY":
                tax_payable = payslip['amount']

            if emp_id not in employee_dict: #check the id for repetation
                employee_dict[emp_id] = {
                    "id": payslip['id'],
                    "staff_id": payslip['staff_id'] or '',
                    "name": payslip['emp_name'] or '',
                    "tax_payer_number":payslip['tax_payer_number'] or '',
                    "wage":payslip['wage']
                }
            employee_data = employee_dict[emp_id] # {id:value, staff_id: value}

            employee_data["taxable_amount"] = employee_data.get("taxable_amount", 0) + taxable_amount # if taxable_amount += taxable_amount else taxable_amount
            employee_data["tax_payable"] = employee_data.get("tax_payable", 0) + tax_payable # if tax_payable  += tax_payable else tax_payable

        employees_data = list(employee_dict.values()) #[{id:value,etc...},etc..]

        file_content = self._generate_employee_paye_report(employees_data)
        
        self.file_name = f"Employee PAYE Report-{fields.Date.today()}.xlsx"
        self.file_data = base64.b64encode(file_content)
        
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model={self._name}&id={self.id}&field=file_data&filename={self.file_name}&download=true",
            'target': 'self',
        }

    def _generate_employee_paye_report(self, employees):
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Paye Report"

        main_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        main_header_font = Font(bold=True, size=11, color="000000")

        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25

        ws['A1'].value = "S/N"
        ws['A1'].fill = main_header_fill
        ws['A1'].font = main_header_font
        ws['A1'].alignment = left_alignment
        ws['A1'].border = thin_border

        ws['B1'].value = "Staff ID"
        ws['B1'].fill = main_header_fill
        ws['B1'].font = main_header_font
        ws['B1'].alignment = center_alignment
        ws['B1'].border = thin_border

        ws['C1'].value = "Name"
        ws['C1'].fill = main_header_fill
        ws['C1'].font = main_header_font
        ws['C1'].alignment = left_alignment
        ws['C1'].border = thin_border

        ws['D1'].value = "Tax Payer Number"
        ws['D1'].fill = main_header_fill
        ws['D1'].font = main_header_font
        ws['D1'].alignment = center_alignment
        ws['D1'].border = thin_border

        ws['E1'].value = "Staff Gross Pay"
        ws['E1'].fill = main_header_fill
        ws['E1'].font = main_header_font
        ws['E1'].alignment = center_alignment
        ws['E1'].border = thin_border

        ws['F1'].value = "Taxable Amount"
        ws['F1'].fill = main_header_fill
        ws['F1'].font = main_header_font
        ws['F1'].alignment = center_alignment
        ws['F1'].border = thin_border

        ws['G1'].value = "Tax Payable"
        ws['G1'].fill = main_header_fill
        ws['G1'].font = main_header_font
        ws['G1'].alignment = center_alignment
        ws['G1'].border = thin_border

        cnt = 1
        current_row = 2
        for emp in employees:
            cell_a = ws.cell(row=current_row, column=1)
            cell_a.value = cnt
            cell_a.alignment = right_alignment
            cell_a.border = thin_border
            
            cell_b = ws.cell(row=current_row, column=2)
            cell_b.value = emp["staff_id"]
            cell_b.alignment = left_alignment
            cell_b.border = thin_border
            
            cell_c = ws.cell(row=current_row, column=3)
            cell_c.value = emp["name"]
            cell_c.alignment = left_alignment
            cell_c.border = thin_border
            
            cell_d = ws.cell(row=current_row, column=4)
            cell_d.value = emp["tax_payer_number"]
            cell_d.alignment = left_alignment
            cell_d.border = thin_border
            
            cell_e = ws.cell(row=current_row, column=5)
            cell_e.value = emp['wage'] or ''
            cell_e.alignment = right_alignment
            cell_e.border = thin_border
            
            cell_f = ws.cell(row=current_row, column=6)
            cell_f.value = emp["taxable_amount"]
            cell_f.alignment = right_alignment
            cell_f.border = thin_border
            
            cell_g = ws.cell(row=current_row, column=7)
            cell_g.value = emp["tax_payable"]
            cell_g.alignment = right_alignment
            cell_g.border = thin_border
            
            cnt += 1
            current_row += 1


        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        file_content = excel_file.read()
        excel_file.close()
        return file_content

    #PFA PENSION REPORT
    def pfa_pension_report(self):
        if self.start_date and self.end_date:
            query = """
                select hr_pay.id as id, hr_emp.id as emp_id, hr_emp.staff_id as staff_id, hr_emp.name as emp_name, 
                hr_emp.rsa_pin as rsa_pin,hr_emp.pfa_name as pfa_name, hr_sal_rul.code as code, hr_pay_line.amount as amount, hr_con.employee_voluntary as employee_voluntary,
                hr_con.employer_voluntary as employer_voluntary from hr_payslip as hr_pay 
                left join hr_employee as hr_emp on hr_pay.employee_id = hr_emp.id
                left join hr_payslip_line as hr_pay_line on hr_pay.id = hr_pay_line.slip_id
                left join hr_salary_rule as hr_sal_rul on hr_pay_line.salary_rule_id = hr_sal_rul.id
                left join hr_contract as hr_con on hr_emp.id = hr_con.employee_id and hr_con.state = 'open'
                where hr_pay.state = 'done' and hr_pay.date_from >= %s and hr_pay.date_to <= %s
            """
            params = (self.start_date, self.end_date)
        else:
            query = """
                select hr_pay.id as id, hr_emp.id as emp_id, hr_emp.staff_id as staff_id, hr_emp.name as emp_name, 
                hr_emp.rsa_pin as rsa_pin,hr_emp.pfa_name as pfa_name, hr_sal_rul.code as code, hr_pay_line.amount as amount, hr_con.employee_voluntary as employee_voluntary,
                hr_con.employer_voluntary as employer_voluntary from hr_payslip as hr_pay 
                left join hr_employee as hr_emp on hr_pay.employee_id = hr_emp.id
                left join hr_payslip_line as hr_pay_line on hr_pay.id = hr_pay_line.slip_id
                left join hr_salary_rule as hr_sal_rul on hr_pay_line.salary_rule_id = hr_sal_rul.id
                left join hr_contract as hr_con on hr_emp.id = hr_con.employee_id and hr_con.state = 'open'
                where hr_pay.state = 'done'
            """
            params = ()

        self.env.cr.execute(query, params)
        hr_payslips = self.env.cr.dictfetchall()

        employee_dict = {}

        for payslip in hr_payslips: 
            emp_id = payslip['emp_id']

            if emp_id not in employee_dict: #check the id for repetation
                employee_dict[emp_id] = {
                    "id": payslip['id'],
                    "staff_id": payslip['staff_id'] or '',
                    "name": payslip['emp_name'] or '',
                    "rsa_pin": payslip['rsa_pin'] or '',
                    "pfa_name": payslip['pfa_name'] or '',
                    "employer_voluntary": payslip['employer_voluntary'] or '',
                    "employee_voluntary": payslip['employee_voluntary'] or '',
                    "employer_pension":0,
                    "employee_pension":0
                }
            employee_data = employee_dict[emp_id] # {id:value, staff_id: value}

            if payslip['code'] == "EMP":
                employee_data["employee_pension"] += payslip['amount']

            if payslip['code'] == 'EMYP':
                employee_data["employer_pension"] += payslip['amount']

        employees_data = list(employee_dict.values()) #[{id:value,etc...},etc..]    

        file_content = self._generate_pfa_pension_report(employees_data)

        self.file_name = f"PFA Pension report-{fields.Date.today()}.xlsx"
        self.file_data = base64.b64encode(file_content)
        
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model={self._name}&id={self.id}&field=file_data&filename={self.file_name}&download=true",
            'target': 'self',
        }

    def _generate_pfa_pension_report(self,employees):
        wb = Workbook()
        ws = wb.active
        ws.title = "PFA Pension Report"

        main_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        main_header_font = Font(bold=True, size=11, color="000000")

        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        ws['A1'].value = "S/N"
        ws['A1'].fill = main_header_fill
        ws['A1'].font = main_header_font
        ws['A1'].alignment = left_alignment
        ws['A1'].border = thin_border

        ws['B1'].value = "Staff ID"
        ws['B1'].fill = main_header_fill
        ws['B1'].font = main_header_font
        ws['B1'].alignment = center_alignment
        ws['B1'].border = thin_border

        ws['C1'].value = "Employee Name"
        ws['C1'].fill = main_header_fill
        ws['C1'].font = main_header_font
        ws['C1'].alignment = left_alignment
        ws['C1'].border = thin_border

        ws['D1'].value = "RSA Pin"
        ws['D1'].fill = main_header_fill
        ws['D1'].font = main_header_font
        ws['D1'].alignment = center_alignment
        ws['D1'].border = thin_border

        ws['E1'].value = "PFA Code/Name"
        ws['E1'].fill = main_header_fill
        ws['E1'].font = main_header_font
        ws['E1'].alignment = center_alignment
        ws['E1'].border = thin_border

        ws['F1'].value = "Employer Contribution"
        ws['F1'].fill = main_header_fill
        ws['F1'].font = main_header_font
        ws['F1'].alignment = center_alignment
        ws['F1'].border = thin_border

        ws['G1'].value = "Employee Contribution"
        ws['G1'].fill = main_header_fill
        ws['G1'].font = main_header_font
        ws['G1'].alignment = center_alignment
        ws['G1'].border = thin_border

        ws['H1'].value = "Employer Voluntary"
        ws['H1'].fill = main_header_fill
        ws['H1'].font = main_header_font
        ws['H1'].alignment = center_alignment
        ws['H1'].border = thin_border

        ws['I1'].value = "Employee Voluntary"
        ws['I1'].fill = main_header_fill
        ws['I1'].font = main_header_font
        ws['I1'].alignment = center_alignment
        ws['I1'].border = thin_border

        ws['J1'].value = "Total Contribution"
        ws['J1'].fill = main_header_fill
        ws['J1'].font = main_header_font
        ws['J1'].alignment = center_alignment
        ws['J1'].border = thin_border

        ws['K1'].value = "Period"
        ws['K1'].fill = main_header_fill
        ws['K1'].font = main_header_font
        ws['K1'].alignment = center_alignment
        ws['K1'].border = thin_border

        cnt = 1
        current_row = 2
        for emp in employees:
            cell_a = ws.cell(row=current_row, column=1)
            cell_a.value = cnt
            cell_a.alignment = right_alignment
            cell_a.border = thin_border
            
            cell_b = ws.cell(row=current_row, column=2)
            cell_b.value = emp['staff_id'] or ''
            cell_b.alignment = left_alignment
            cell_b.border = thin_border
            
            cell_c = ws.cell(row=current_row, column=3)
            cell_c.value = emp['name']
            cell_c.alignment = left_alignment
            cell_c.border = thin_border
            
            cell_d = ws.cell(row=current_row, column=4)
            cell_d.value = emp['rsa_pin'] or ''
            cell_d.alignment = left_alignment
            cell_d.border = thin_border
            
            cell_e = ws.cell(row=current_row, column=5)
            cell_e.value = emp['pfa_name'] or ''
            cell_e.alignment = left_alignment
            cell_e.border = thin_border

            total_contribution = 0

            cell_f = ws.cell(row=current_row, column=6)
            cell_f.value = emp['employer_pension'] or ''
            total_contribution += int(emp['employer_pension'])
            cell_f.alignment = left_alignment
            cell_f.border = thin_border

            cell_g = ws.cell(row=current_row, column=7)
            cell_g.value = emp['employee_pension'] or ''
            total_contribution += int(emp['employee_pension'])
            cell_g.alignment = left_alignment
            cell_g.border = thin_border


            cell_h = ws.cell(row=current_row, column=8)
            cell_h.value = emp['employer_voluntary'] or ''
            cell_h.alignment = left_alignment
            cell_h.border = thin_border

            cell_i = ws.cell(row=current_row, column=9)
            cell_i.value = emp['employee_voluntary'] or ''
            cell_i.alignment = left_alignment
            cell_i.border = thin_border

            employer_vol = emp.get('employer_voluntary') or 0
            employee_vol = emp.get('employee_voluntary') or 0
            total_contribution += int(employer_vol) + int(employee_vol)

            cell_j = ws.cell(row=current_row, column=10)
            cell_j.value = total_contribution or ''
            cell_j.alignment = left_alignment
            cell_j.border = thin_border

            period = f"{self.start_date.strftime('%m/%y')} - {self.end_date.strftime('%m/%y')}" if self.start_date and self.end_date else ""    

            cell_k = ws.cell(row=current_row, column=11)
            cell_k.value = period or ''
            cell_k.alignment = left_alignment
            cell_k.border = thin_border

            cnt += 1
            current_row += 1
            

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        file_content = excel_file.read()
        excel_file.close()
        return file_content


    #INVENTORY  HELD REPORT
    def inventory_held_data(self):
        domain = [('state', '=', 'assigned'),('sale_id.state', '=', 'sale'),('sale_id.date_order','>=',self.start_date),('sale_id.date_order','<=',self.end_date)] if self.start_date and self.end_date else [('state', '=', 'assigned'),('sale_id.state', '=', 'sale')]
        deliveries = self.env['stock.picking'].search(domain)

        product_ids_demand = {}     # {product_id: {'demand': qty}}
        product_ids_available = {}  # {product_id: {'available': qty}}
        customers_dict = {}         # {product_id: {customer: count}}

        for delivery in deliveries:
            customer = delivery.sale_id.partner_id.name or ''
            for move in delivery.move_ids_without_package:

                product_id = move.product_id.id
                if not product_id:
                    continue

                #  delivery quantity field
                booked_qty = move.product_uom_qty    

                if booked_qty <= 0:
                    continue

                #  Initialize dict for each product
                if product_id not in product_ids_demand:
                    product_ids_demand[product_id] = {"demand": 0}
                    product_ids_available[product_id] = {"available": 0}
                    customers_dict[product_id] = {}

                #  Add booked qty
                product_ids_demand[product_id]['demand'] += booked_qty

                #  Add customer
                customers_dict[product_id].setdefault(customer, 0)
                customers_dict[product_id][customer] += 1

        #  Compute qty_available from stock
        product_model = self.env['product.product']
        for product_id in product_ids_demand.keys():
            product = product_model.browse(product_id)
            product_ids_available[product_id]['available'] = int(product.qty_available)

        #  Generate Excel
        file_content = self._generate_inventory_held_report(
            product_ids_demand,
            product_ids_available,
            customers_dict
        )

        self.file_name = f"Custom Report that shows inventory held for a client(Client Stock)-{fields.Date.today()}.xlsx"
        self.file_data = base64.b64encode(file_content)
        
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model={self._name}&id={self.id}&field=file_data&filename={self.file_name}&download=true",
            'target': 'self',
        }

    def _generate_inventory_held_report(self, product_ids_demand, product_ids_available, customers_dict):

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Held For A Client Report"

        # ---------- Styles ----------
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ---------- Main Title ----------
        ws.merge_cells('A1:F1')
        ws['A1'].value = "Custom Report that shows inventory held for a client (Client Stock)"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = center

        # ---------- Header Row ----------
        headers = [
            "Product Name",
            "Product Description",
            "Quantity Available in Stock",
            "Booked for Client",
            "Net Balance",
            "Client's Name"
        ]

        col = 1
        for h in headers:
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = bold
            cell.alignment = center
            cell.border = border
            col += 1

        # ---------- Prepare Product Lines ----------
        product_obj = self.env["product.product"]

        row = 4

        for product_id in product_ids_demand.keys():

            product = product_obj.browse(product_id)
            
            description = product.description_sale or ""
            qty_available = int(product_ids_available[product_id]['available'])
            booked_qty = int(product_ids_demand[product_id]['demand'])
            net_balance = qty_available - booked_qty
            display_name = f"[{product.default_code}] - {product.name}" if product.default_code else product.name

            # Prepare client list
            customers = []
            for cust in customers_dict.get(product_id, {}).keys():
                customers.append(cust)

            customers_txt = "/".join(customers)

            # ---------- Write row ----------
            data = [
                display_name,
                description,
                qty_available,
                booked_qty,
                net_balance,
                customers_txt
            ]

            for i, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=i, value=value)
                cell.border = border
                cell.alignment = center if i <= 2 else center

                # Red for negative or booked > 0
                if i == 4 and booked_qty > 0:
                    cell.font = Font(color="FF0000")
                if i == 5 and net_balance < 0:
                    cell.font = Font(color="FF0000")

            row += 1

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Give padding of +2 for nicer spacing
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        # ---------- Generate File ----------
        file_buf = io.BytesIO()
        wb.save(file_buf)
        file_buf.seek(0)
        return file_buf.read()

    #CLIENT STOCK REPORT
    def client_stock_data(self):        
        sale_order = self.env['sale.order'].search([('state', '=', 'sale')])
        stock_picking = self.env['stock.picking']
        product_product = self.env['product.product']

        product_ids_demand = dict() # {product_id: {demand:value}}
        product_ids_available = dict() # {product_id: {available:value}}
        deliveries = []
        customers_dict = dict()  # {product_id: {customer_name: count}}

        for record in sale_order:
            record_date = record.date_order.date()
            if self.start_date and self.end_date:
                if record_date >= self.start_date and record_date <= self.end_date:
                    delivery = stock_picking.search([('origin', '=', record.name), ('state', '!=', 'done')], limit=1)
                if delivery:
                    deliveries.append(delivery)
            else:
                delivery = stock_picking.search([('origin', '=', record.name), ('state', '!=', 'done')], limit=1)
                if delivery:
                    deliveries.append(delivery)

        for delivery in deliveries:
            customer_name = delivery.sale_id.partner_id.name
            for move in delivery.move_ids:
                product_id = move.product_id.id
                if product_id in product_ids_demand:
                    product_ids_demand[product_id]['demand'] += move.product_uom_qty
                else:
                    product_ids_demand[product_id] = {'demand': move.product_uom_qty}
                    product_ids_available[product_id] = {'available': 0}
                    customers_dict[product_id] = {}

                if customer_name in customers_dict[product_id]:
                    customers_dict[product_id][customer_name] += 1
                else:
                    customers_dict[product_id][customer_name] = 1

        for product_id in product_ids_available.keys():
            product = product_product.browse(product_id)
            product_ids_available[product_id]['available'] = int(product.qty_available)

        file_content = self._generate_client_stock_report(product_ids_demand, product_ids_available, customers_dict)

        self.file_data = base64.b64encode(file_content)

        self.file_name = f"Client_Stock_Report_{fields.Date.today()}.xlsx"

        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model={self._name}&id={self.id}&field=file_data&filename={self.file_name}&download=true",
            'target': 'self',
        }

    def _generate_client_stock_report(self, product_ids_demand, product_ids_available, customers_dict):
        wb = Workbook()
        ws = wb.active
        ws.title = "Client Stock Report"

        main_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        main_header_font = Font(bold=True, size=11, color="4169E1")
        
        column_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        column_header_font = Font(bold=True, size=10)
        
        sub_header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        sub_header_font = Font(bold=True, size=10, color="4169E1")
        
        category_fill = PatternFill(start_color="6699FF", end_color="6699FF", fill_type="solid")
        category_font = Font(bold=True, color="FF0000", size=10)
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 60

        main_header = ws['A1']
        main_header.value = "CLIENTS' STOCK LIST"
        main_header.fill = main_header_fill
        main_header.font = main_header_font
        main_header.alignment = center_alignment
        main_header.border = thin_border

        ws['A2'].value = "Product Name"
        ws['A2'].fill = column_header_fill
        ws['A2'].font = column_header_font
        ws['A2'].alignment = center_alignment
        ws['A2'].border = thin_border
        
        ws['B2'].value = ""
        ws['B2'].fill = column_header_fill
        ws['B2'].border = thin_border
        
        ws['C2'].value = "Quantity"
        ws['C2'].fill = column_header_fill
        ws['C2'].font = column_header_font
        ws['C2'].alignment = center_alignment
        ws['C2'].border = thin_border
        
        ws['F2'].value = "Client's Name"
        ws['F2'].fill = column_header_fill
        ws['F2'].font = column_header_font
        ws['F2'].alignment = center_alignment
        ws['F2'].border = thin_border

        ws['A3'].value = ""
        ws['A3'].border = thin_border
        ws['B3'].value = ""
        ws['B3'].border = thin_border
        
        ws['C3'].value = "Owed"
        ws['C3'].fill = sub_header_fill
        ws['C3'].font = sub_header_font
        ws['C3'].alignment = center_alignment
        ws['C3'].border = thin_border
        
        ws['D3'].value = "Avail"
        ws['D3'].fill = sub_header_fill
        ws['D3'].font = sub_header_font
        ws['D3'].alignment = center_alignment
        ws['D3'].border = thin_border
        
        ws['E3'].value = "Net Balance"
        ws['E3'].fill = sub_header_fill
        ws['E3'].font = sub_header_font
        ws['E3'].alignment = center_alignment
        ws['E3'].border = thin_border
        
        ws['F3'].value = ""
        ws['F3'].border = thin_border
        
        ws['G2'].value = "Delivery Order"
        ws['G2'].fill = column_header_fill
        ws['G2'].font = column_header_font
        ws['G2'].alignment = center_alignment
        ws['G2'].border = thin_border

        total_owed = 0
        total_available = 0
        total_net_balance = 0

        product_obj = self.env['product.product']
        products_by_category = {}
        
        for product_id in product_ids_demand.keys():
            product = product_obj.browse(product_id)
            category = product.categ_id.name if product.categ_id else 'Uncategorized'
            
            if category not in products_by_category:
                products_by_category[category] = []

            customer_list = []
            for customer_name, count in customers_dict.get(product_id, {}).items():
                if count > 1:
                    customer_list.append(f"{customer_name}({count})")
                else:
                    customer_list.append(customer_name)
            
            delivery_order = []
            pickings = self.env['stock.picking'].search([       
                ('sale_id.state', '=', 'sale'),
                ('product_id','=',product_id),
                ('state','!=','done')])
            for picking in pickings:
                delivery_order.append(picking.name)

            display_name = f"[{product.default_code}] - {product.name}" if product.default_code else product.name

            products_by_category[category].append({
                'product': product,
                'default_code': product.default_code or "",
                'name': display_name,
                'product_id': product_id,
                'demand': product_ids_demand[product_id]['demand'],
                'available': product_ids_available[product_id]['available'],
                'delivery_orders': ','.join(delivery_order),
                'customers': ', '.join(customer_list)
            })

        
        current_row = 4
        
        for category in sorted(products_by_category.keys()):
            ws.merge_cells(f'A{current_row}:G{current_row}')
            category_cell = ws[f'A{current_row}']
            category_cell.value = category.upper()
            category_cell.fill = category_fill
            category_cell.font = category_font
            category_cell.alignment = left_alignment
            category_cell.border = thin_border
            current_row += 1

            for item in sorted(products_by_category[category], key=lambda x: x['default_code']):
                demand = int(item['demand'])
                available = int(item['available'])
                net_balance = available - demand  # Net Balance = Available - Demand

                total_owed += demand
                total_available += available
                total_net_balance += net_balance
    
                cell_a = ws.cell(row=current_row, column=1)
                cell_a.value = item['name']
                cell_a.alignment = left_alignment
                cell_a.border = thin_border
                
                cell_c = ws.cell(row=current_row, column=3)
                cell_c.value = demand
                cell_c.alignment = left_alignment
                cell_c.border = thin_border

                if demand > 0:
                    cell_c.font = Font(color="FF0000")
                
                cell_d = ws.cell(row=current_row, column=4)
                cell_d.value = available
                cell_d.alignment = center_alignment
                cell_d.border = thin_border
                
                cell_e = ws.cell(row=current_row, column=5)
                cell_e.value = net_balance
                cell_e.alignment = center_alignment
                cell_e.border = thin_border

                if net_balance < 0:
                    cell_e.font = Font(color="FF0000")
                
                cell_f = ws.cell(row=current_row, column=6)
                cell_f.value = item['customers']
                cell_f.alignment = center_alignment
                cell_f.border = thin_border

                cell_g = ws.cell(row=current_row, column=7)
                cell_g.value = item['delivery_orders']
                cell_g.alignment = center_alignment
                cell_g.border = thin_border
                cell_g.font = Font(color="FF0000")
                
                current_row += 1
            current_row += 1

        ws['C1'].value = total_owed
        ws['C1'].fill = main_header_fill
        ws['C1'].font = Font(bold=True, size=10, color="000000")
        ws['C1'].alignment = center_alignment
        ws['C1'].border = thin_border

        ws['D1'].value = total_available
        ws['D1'].fill = main_header_fill
        ws['D1'].font = Font(bold=True, size=10, color="000000")
        ws['D1'].alignment = center_alignment
        ws['D1'].border = thin_border

        ws['E1'].value = total_net_balance
        ws['E1'].fill = main_header_fill
        ws['E1'].font = main_header_font
        ws['E1'].alignment = center_alignment
        ws['E1'].border = thin_border

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        file_content = excel_file.read()
        excel_file.close()

        return file_content

    # PAYROLL SCHEDULE REPORT 
    def _generate_payroll_schedule_report(self):
        if self.start_date and self.end_date:
            query = """
                select hr_pay.id as id, hr_emp.id as emp_id,hr_dep.name as dep_name, hr_emp.staff_id as staff_id, hr_emp.name as emp_name, 
                hr_pay_line.name as col_name,hr_pay_line.code as code, hr_pay_line.amount as amount
                from hr_payslip as hr_pay 
                left join hr_employee as hr_emp on hr_pay.employee_id = hr_emp.id
                full join hr_department as hr_dep on hr_emp.department_id = hr_dep.id 
                left join hr_payslip_line as hr_pay_line on hr_pay.id = hr_pay_line.slip_id
                where hr_pay.state = 'done' and hr_pay.date_from >= %s and hr_pay.date_to <= %s
            """
            params = (self.start_date, self.end_date)
        else:
            query = """
                select hr_pay.id as id, hr_emp.id as emp_id, hr_dep.name as dep_name,hr_emp.staff_id as staff_id, hr_emp.name as emp_name, 
                hr_pay_line.name as col_name,hr_pay_line.code as code, hr_pay_line.amount as amount
                from hr_payslip as hr_pay 
                left join hr_employee as hr_emp on hr_pay.employee_id = hr_emp.id
                full join hr_department as hr_dep on hr_emp.department_id = hr_dep.id
                left join hr_payslip_line as hr_pay_line on hr_pay.id = hr_pay_line.slip_id
                where hr_pay.state = 'done'
            """
            params = ()

        self.env.cr.execute(query, params)
        hr_payslips = self.env.cr.dictfetchall()

        columns = set()

        for data in hr_payslips:
            if data['code'] == "LEAVE":
                continue
            
            columns.add(data['col_name'])

        columns = sorted(columns)

        employees = []
        employee_dict = {} # {emp_id(1): {id:value, staff_id:value, etc..}}

        for payslip in hr_payslips:
            emp_id = payslip['emp_id']

            if emp_id not in employee_dict: #check the id for repetation
                employee_dict[emp_id] = {
                    "id": emp_id,
                    "staff_id": payslip['staff_id'],
                    "name": payslip['emp_name'],
                    "department": payslip['dep_name']['en_US']
                }
            
            employee_data = employee_dict[emp_id] # {id:value, staff_id: value}
            amount = payslip['amount']

            employee_data[payslip['col_name']] = employee_data.get(payslip['col_name'],0) + amount 
            
        total = {}
        for col in columns:
            total[col] = 0

        employees = list(employee_dict.values()) #[{id:value,etc...},etc..]

        for emp in employees:
            for key, val in emp.items():
                if key in total:
                    total[key] += val
         
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        writer.writerow(["S/N", "Staff ID", "Employee Name", "Department", *columns])

        cnt = 1
        for emp in employees:
            row = [
                cnt,
                emp.get("staff_id", ""),
                emp.get("name", ""),
                emp.get("department", ""),
            ]
            
            for col in columns:
                row.append(emp.get(col,0))
            
            writer.writerow(row)
            cnt += 1
        last_row = [cnt,"","Total","",*total.values()]
        writer.writerow(last_row)

        csv_data = buffer.getvalue().encode()
        buffer.close()


        self.file_name = f"Payroll Schedule Report-{fields.Date.today()}.csv"
        
        self.file_data = base64.b64encode(csv_data)
        
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model={self._name}&id={self.id}&field=file_data&filename={self.file_name}&download=true",
            'target': 'self',
        }
