from odoo import fields, models, api
from odoo.exceptions import UserError
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class HrEmployee(models.Model):
    _inherit = "hr.employee"

    tax_payer_number = fields.Char("Tax Payer Number")
    staff_id = fields.Char("Staff ID")
    rsa_pin = fields.Char("RSA PIN")
    pfa_name = fields.Char("PFA Code/Name")
    file_data = fields.Binary("File",store=True, attachment=False)
    file_name = fields.Char("File Name")

    @api.model
    def action_payment_schedule_report(self):
        query = """
            select hr_emp.id,hr_emp.staff_id,hr_emp.name as emp_name,hr_dep.name as dep_name,r_p_b.acc_number,r_b.name as bank_name,hr_pay_line.name as col_name, hr_pay_line.amount as amount
            from hr_employee as hr_emp 
            left join hr_department as hr_dep on hr_emp.department_id = hr_dep.id 
            left join res_partner_bank as r_p_b on r_p_b.id = hr_emp.bank_account_id
            left join res_bank as r_b on r_b.id = r_p_b.bank_id
            left join hr_payslip as hr_pay on hr_emp.id = hr_pay.employee_id
            left join hr_payslip_line as hr_pay_line on hr_pay.id = hr_pay_line.slip_id
        """
        self.env.cr.execute(query)
        employees_data = self.env.cr.dictfetchall()
       
        employees = {}
        for emp in employees_data:
            if  emp['id'] not in employees:
                employees[emp['id']] = {
                    'staff_id' : emp['staff_id'],
                    'emp_name' : emp['emp_name'],
                    'dep_name' : emp['dep_name'],
                    'acc_number' : emp['acc_number'],
                    'bank_name' : emp['bank_name'],
                    'net' : 0
                }
            else:
                if emp['col_name'] == "Net Salary":
                    employees[emp['id']]['net'] +=  emp['amount'] 
                
            
        file_content = self._generate_payment_report(employees)

        file_name = f"Payment Schedule Report-{fields.Date.today()}.xlsx"
        file_data = base64.b64encode(file_content)
        
        employee_id = False
        for emp in employees_data:
            if emp['id']:
                employee_id = emp['id']
                break

        if employee_id:
            update_query = """
                update hr_employee 
                set file_data = %s, file_name = %s
                where id = %s
            """
            self.env.cr.execute(update_query, (file_data, file_name, employee_id))
            self.env.cr.commit()
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model=hr.employee&id={employee_id}&field=file_data&filename={file_name}&download=true",
            'target': 'self',
        }

    def _generate_payment_report(self, employees):
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees Details"

        main_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        main_header_font = Font(bold=True, size=11, color="000000")

        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        right_alignment = Alignment(horizontal='right', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25

        ws['A1'].value = "S/N"
        ws['A1'].fill = main_header_fill
        ws['A1'].font = main_header_font
        ws['A1'].alignment = left_alignment
        ws['A1'].border = thin_border

        ws['B1'].value = "Staff ID"
        ws['B1'].fill = main_header_fill
        ws['B1'].font = main_header_font
        ws['B1'].alignment = center_alignment
        ws['B1'].border = thin_border

        ws['C1'].value = "Name"
        ws['C1'].fill = main_header_fill
        ws['C1'].font = main_header_font
        ws['C1'].alignment = left_alignment
        ws['C1'].border = thin_border

        ws['D1'].value = "Department"
        ws['D1'].fill = main_header_fill
        ws['D1'].font = main_header_font
        ws['D1'].alignment = center_alignment
        ws['D1'].border = thin_border

        ws['E1'].value = "Bank Name"
        ws['E1'].fill = main_header_fill
        ws['E1'].font = main_header_font
        ws['E1'].alignment = left_alignment
        ws['E1'].border = thin_border

        ws['F1'].value = "Account No"
        ws['F1'].fill = main_header_fill
        ws['F1'].font = main_header_font
        ws['F1'].alignment = left_alignment
        ws['F1'].border = thin_border

        ws['G1'].value = "Net Pay"
        ws['G1'].fill = main_header_fill
        ws['G1'].font = main_header_font
        ws['G1'].alignment = left_alignment
        ws['G1'].border = thin_border

        cnt = 1
        current_row = 2
        total_net_pay = 0

        for key,value in employees.items():
            cell_a = ws.cell(row=current_row, column=1)
            cell_a.value = cnt
            cell_a.alignment = right_alignment
            cell_a.border = thin_border
            
            cell_b = ws.cell(row=current_row, column=2)
            cell_b.value = value['staff_id'] 
            cell_b.alignment = left_alignment
            cell_b.border = thin_border
            
            cell_c = ws.cell(row=current_row, column=3)
            cell_c.value = value['emp_name'] 
            cell_c.alignment = left_alignment
            cell_c.border = thin_border

            cell_d = ws.cell(row=current_row, column=4)
            cell_d.value = value['dep_name']['en_US'] if value['dep_name'] else '' 
            cell_d.alignment = left_alignment
            cell_d.border = thin_border

            cell_e = ws.cell(row=current_row, column=5)
            cell_e.value = value['bank_name'] 
            cell_e.alignment = right_alignment
            cell_e.border = thin_border
            
            cell_f = ws.cell(row=current_row, column=6)
            cell_f.value = value['acc_number'] 
            cell_f.alignment = right_alignment
            cell_f.border = thin_border
            
            cell_g = ws.cell(row=current_row, column=7)
            cell_g.value = value['net']
            cell_g.alignment = right_alignment
            cell_g.border = thin_border

            total_net_pay += cell_g.value
            
            cnt += 1
            current_row += 1
        
        last_row_a = ws.cell(row=current_row, column=1)
        last_row_a.value = cnt
        last_row_a.alignment = right_alignment
        last_row_a.border = thin_border
        last_row_a.fill = main_header_fill
        last_row_a.font = main_header_font

        last_row_b = ws.cell(row=current_row, column=2)
        last_row_b.border = thin_border

        last_row_c = ws.cell(row=current_row, column=3)
        last_row_c.value = "Total"
        last_row_c.alignment = left_alignment
        last_row_c.border = thin_border
        last_row_c.fill = main_header_fill
        last_row_c.font = main_header_font

        last_row_d = ws.cell(row=current_row, column=4)
        last_row_d.border = thin_border

        last_row_e = ws.cell(row=current_row, column=5)
        last_row_e.border = thin_border

        last_row_f = ws.cell(row=current_row, column=6)
        last_row_f.border = thin_border

        last_row_g = ws.cell(row=current_row, column=7)
        last_row_g.value = total_net_pay
        last_row_g.alignment = right_alignment
        last_row_g.border = thin_border
        last_row_g.fill = main_header_fill
        last_row_g.font = main_header_font

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        file_content = excel_file.read()
        excel_file.close()
        return file_content