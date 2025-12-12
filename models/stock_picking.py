# -*- coding: utf-8 -*-

from odoo import models,api,fields
from odoo.exceptions import UserError
import io
import base64
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class StockPicking(models.Model):
    _inherit = "stock.picking"
    file_data = fields.Binary("File")
    file_name = fields.Char("File Name")

    @api.model
    def action_warehouse_data(self):
        sale_orders = self.env["sale.order"].search([('state', '=', 'sale')])
        pickings = self.env["stock.picking"].search([
        ('sale_id', 'in', sale_orders.ids),
        ('state', '!=', 'done'),
        ])
        product_data = {}
        warehouse_list = self.env["stock.warehouse"].search([], order="sequence asc")

        for picking in pickings:
            for move in picking.move_ids_without_package:
                product = move.product_id.product_tmpl_id
                if not product:
                    continue

                done_qty = sum(move.move_line_ids.mapped("qty_done"))
                owed_qty = move.product_uom_qty - done_qty

                if owed_qty <= 0:
                    continue

                pid = product.id
                if pid not in product_data:
                    product_data[pid] = {
                        "product": product,
                        "description": product.description_sale or "",
                        "product_name":product.name or "",
                        "default_code": product.default_code or "",
                        "delivery_order": set(),
                        "client_stock": 0,
                        "warehouses": {w.id: 0 for w in warehouse_list},
                    }

                product_data[pid]["client_stock"] += owed_qty
    
                delivery_order = picking.name
                if delivery_order:
                    product_data[pid]["delivery_order"].add(delivery_order)

        for pid, pdata in product_data.items():
            product = pdata["product"]
            for wh in warehouse_list:
                domain = [
                    ('product_id.product_tmpl_id', '=', pid),
                    ('location_id', '=', wh.lot_stock_id.id),
                ]
                quants = self.env["stock.quant"].search(domain)
                qty = sum(quants.mapped('quantity'))
                pdata["warehouses"][wh.id] = qty
                    
        file_data = self._generate_custom_inventory_report(product_data, warehouse_list)
    
        file_name = f"Custom Inventory Report By Warehouse-{fields.Date.today()}.xlsx"
        file_data = base64.b64encode(file_data)

        id = False
        for data in pickings:
            id = data.id
            data.file_name = file_name
            data.file_data = file_data
        
        return {
            'type': 'ir.actions.act_url',
            'url': f"/web/content/?model={self._name}&id={id}&field=file_data&filename={file_name}&download=true",
            'target': 'self',
        }

    def _generate_custom_inventory_report(self, product_data, warehouse_list):

        wb = Workbook()
        ws = wb.active
        ws.title = "Custom Inventory  Report By Warehouse"

        # Styles
        header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        header_font = Font(color="000000", bold=True)
        center = Alignment(horizontal="center")
        left = Alignment(horizontal="left")
        bold = Font(bold=True)
        border = Border(left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"))
    
        total_columns = 3 + len(warehouse_list) + 2

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        block_cell = ws.cell(row=1, column=1, value="Custom Inventory report by warehouse showing.")
        block_cell.font = Font(bold=False, size=14)
        block_cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_columns)
        block_cell = ws.cell(row=3, column=1, value="STOCK BALANCE")
        block_cell.font = Font(bold=True, size=14)
        block_cell.alignment = Alignment(horizontal="center", vertical="center")

        thick_black = Border(
            left=Side(style="thick", color="000000"),
            right=Side(style="thick", color="000000"),
            top=Side(style="thick", color="000000"),
            bottom=Side(style="thick", color="000000"),
        )

        # Apply the thick border to each merged cell
        for col in range(1, total_columns + 1):
            ws.cell(row=3, column=col).border = thick_black

        headers = [
            "Product Name",
            "Product Description",
            "Delivery Orders",
        ]

        # add warehouse names (in the sorted sequence order)
        headers += [wh.name for wh in warehouse_list]

        # add client stock + net balance at the end
        headers += ["Client Stock", "Net Balance"]

        # Write headers in row 5
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        row = 6
        for pid, pdata in product_data.items():

            # values in EXACT warehouse order
            warehouse_values = [pdata["warehouses"][w.id] for w in warehouse_list]

            delivery_orders = ",".join(sorted(list(pdata["delivery_order"])))
            total_wh = sum(warehouse_values)
            net_balance = total_wh - pdata["client_stock"]

            name = (f"[{pdata['default_code']}] - {pdata['product_name']}" 
                            if pdata['default_code']
                            else pdata['product_name'])
            pdata["name"] = name

            row_data = [
                pdata["name"],
                pdata["description"],
                delivery_orders,
            ]

            # warehouse columns (exact order)
            row_data += warehouse_values

            # client stock + net balance
            row_data += [
                pdata["client_stock"],
                net_balance,
            ]

            # Write data row
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = center

            row += 1

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Give padding of +3 for nicer spacing
            adjusted_width = max_length + 3
            ws.column_dimensions[col_letter].width = adjusted_width

        # Return file
        stream = io.BytesIO()
        wb.save(stream)
        return stream.getvalue()